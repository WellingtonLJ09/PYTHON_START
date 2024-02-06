from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

#1- Lê pasta de trabalho e planilha
wb = load_workbook("data/barchart.xlsx")
sheet = wb["Relatorio"]

#2- Referencias das linhs e colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#3 -Incluindo Fórmula
sheet["B6"] = "=SUM(B2:B5)"
sheet["B6"].style = "Currency"

wb.save("test.xlsx")